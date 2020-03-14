import openpyxl
import os
import pyperclip
import re
# set up file
'''print('Nhập tên file excel cần dò')
exname=input()
print('Nhập tên sheet, để trống sẽ mặc định là Sheet1')
sheetname=input()'''
exname = Xuat.xlsx
sheetname = Sheet1
# set up excel
kihieu = re.compile(r'''
^(\w{1,4}\d{2,4}) # ma so
(\s|,)
(\w{1,7}) # size
(\s|,)
(.*?) # mau
(\s|,)
''', re.VERBOSE)
wb = openpyxl.load_workbook('exname')
sheet = wb.get_sheet_by_name(sheetname)
maxcol = sheet.max_column
# to mau
ft = openpyxl.styles.fonts.Font(
    color=openpyxl.styles.colors.RED, size=14, bold=True)
# giao dien
print('Chú ý: Ctrl C đoạn mã trước khi sử dụng')
print('1.c để thêm')
print('3.` để thoát')
# them hang loat
while True:
    x = input()
    if x == '`':
        break
    elif x == 'c':
        maxcol = sheet.max_column
        letter = pyperclip.paste()
        test = letter.splitlines()
        for z in range(0, len(test)):
            n = kihieu.search(test[z])
            if n == None:
                print('Lỗi ko đọc được thông tin cột thứ '+str(i))
            else:
                maso = n.group(1)
                size = n.group(3)
                mau = n.group(5)
                # xu ly size nhap
                if len(size) == 1:
                    size = 'size'+str(size).capitalize()
                else:
                    size = size[0:4]+str(size[4]).capitalize()
                c = 2
                i = 2
                r = 3
                if maso == None or size == None or mau == None:
                    print('Lỗi 1 hoặc nhiều thông tin bị thiếu ở hàng '+str(i))
                else:
                    while c <= maxcol:
                        if size == sheet.cell(row=1, column=c).value:
                            while c <= maxcol:
                                if mau == str(sheet.cell(row=2, column=c).value).strip():
                                    excol = c
                                    c = maxcol+1
                                    size = ' '
                                else:
                                    c = c+1
                        else:
                            c = c+1
                    while r <= sheet.max_row:
                        if maso == str(sheet.cell(row=r, column=1).value).strip():
                            print('Địa chỉ là ' +
                                  str(sheet.cell(row=r, column=excol).coordinate))
                            sheet.cell(
                                row=r, column=excol).font = ft
                            r = sheet.max_row + 1
# save
print('Ấn 1 để lưu')
print('Ấn 2 để ko lưu')
x = input()
if x == '1':
    print('Nhập tên của file lưu, bỏ trống sẽ định dạng theo tên Tim')
    name = input()
    print('...')
    wb.save('Tim.xlsx')
    print('Đã save với tên Tim.xlsx')
print('Đã thoát')
