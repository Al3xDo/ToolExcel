import openpyxl
import pyperclip
import os
import re
import datetime
# set up file
# tạo và viết cookies
'''dirname = os.path.join(os.getcwd(), 'fileXuat.txt')
if not os.path.exists():
    os.makedirs('dirname')
os.chdir(dirname)'''
'''dirname = os.path.join(os.getcwd(), 'ToolExcel')
os.chdir(dirname)'''
wb = openpyxl.load_workbook('Xuat.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
# viet regex
kihieu = re.compile(r'''
^(\w{1,4}\d{2,4}) # ma so
(\s|,)
(\w{5,7}) # size
(\s|,)
(.*?) # mau
(\s|,)
(\d{1,3}) # so luong
''', re.VERBOSE)
# add nhanh cac cot
# giao dien
print('Lưu ý save excel trước khi dùng')
print('1. c để quét nội dung (can copy vao clipboard truoc)')
print('3.` thoát')
while True:
    # kiem tra dau vao
    print('Mời bạn nhập')
    x = input()
    if x == '`':
        break
    # chuc nang 1
    elif x == 'c':
        maxcol = sheet.max_column
        letter = pyperclip.paste()
        test = letter.splitlines()
        for z in range(0, len(test)):
            n = kihieu.search(test[z])
            if n == None:
                print('Lỗi ko đọc được thông tin cột thứ '+str(i))
            else:
                maso = n.group(1)
                size = n.group(3)
                mau = n.group(5)
                soluong = int(n.group(7))
                size = size[0:4]+str(size[4]).capitalize()
                c = 2
                i = 2
                r = 3
                if maso == None or size == None or mau == None or soluong == None:
                    print('Lỗi 1 hoặc nhiều thông tin bị thiếu ở hàng '+str(i))
                else:
                    while c <= maxcol:
                        if size == sheet.cell(row=1, column=c).value:
                            while c <= maxcol:
                                if mau == str(sheet.cell(row=2, column=c).value).strip():
                                    excol = c
                                    c = maxcol+1
                                    size = ' '
                                else:
                                    c = c+1
                        else:
                            c = c+1
                    while r <= sheet.max_row:
                        if maso == str(sheet.cell(row=r, column=1).value).strip():
                            if sheet.cell(row=r, column=excol).value == None:
                                print('ô ko tồn tại')
                            else:
                                rowmaso = r
                                num = int(sheet.cell(
                                    row=r, column=excol).value)-soluong
                                sheet.cell(row=r, column=excol).value = num
                                r = sheet.max_row+1
                                print('Đã chỉnh xong hàng ở địa chỉ ' +
                                      sheet.cell(row=rowmaso, column=excol).coordinate)
                        else:
                            r = r+1
        print('Hoàn tất')
# chuc nang 2
'''
    elif x == 'z':
        maxcol = sheet.max_column
        test = open('ma.txt')
        test1 = test.readlines()
        for t in test1:
            t = t.strip()
            g = kihieu.search(t)
            if g == None:
                print('Loi')
            else:
                maso = g.group(1)
                size = g.group(3)
                mau = g.group(5)
                soluong = int(g.group(7))
                size = size[0:4]+str(size[4]).capitalize()
                c = 2
                i = 2
                r = 3
                while c <= maxcol:
                    if size == sheet.cell(row=1, column=c).value:
                        while c <= maxcol:
                            if mau == sheet.cell(row=2, column=c).value:
                                excol = c
                                c = maxcol+1
                                size = ' '
                            else:
                                c = c+1
                    else:
                        c = c+1
                while r <= sheet.max_row:
                    if maso == str(sheet.cell(row=r, column=1).value).strip():
                        if sheet.cell(row=r, column=excol).value == None:
                            print('ô ko tồn tại')
                        else:
                            num = int(sheet.cell(
                                row=r, column=excol).value)-soluong
                            sheet.cell(row=r, column=excol).value = num
                            r = sheet.max_row+1
                            print('Đã chỉnh xong')
                    else:
                        r = r+1
        test.close()
        print('Hoàn tất')
'''
# cuoi
print('Ấn 1 để lưu')
print('Ấn 2 để ko lưu')
x = input()
if x == '1':
    print('Nhập tên của file lưu, bỏ trống sẽ định dạng theo tên Xuat Ngay Thang')
    name = input()
    print('...')
    if name == '' or ' ':
        today = datetime.date.today()
        day = today.day
        month = today.month
        name = str(day)+'-'+str(month)
    wb.save('Xuat '+name+' .xlsx')
    print('Đã save với tên Xuat '+name+'.xlsx')
print('Đã hoàn tất')
